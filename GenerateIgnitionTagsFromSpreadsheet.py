#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generate Ignition Tag Import JSON from an Excel workbook.

Expected sheets:
    Settings
    Tags

Expected Settings keys (column A=key, B=value):
    Root Folder
    OPC Server Default
    OPC Item Path Pattern
    Default Value Source
    Default Tag Group
    Fail On Warnings

Expected Tags columns:
    include
    name
    dataType
    documentation
    engUnit
    engLow
    engHigh
    Historize
    alarm
    folderPath

Datatype mapping:
    BOOL -> Boolean
    INT  -> Int4
    REAL -> Float4
"""

import json
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook


DATATYPE_MAP = {
    "BOOL": "Boolean",
    "INT": "Int4",
    "REAL": "Float4"
}


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def bool_from_excel(value, default=False):
    if value is None or value == "":
        return default

    value = str(value).strip().lower()

    return value in (
        "true",
        "1",
        "yes",
        "y",
        "x"
    )


def read_key_value_sheet(workbook, sheet_name):
    result = {}

    if sheet_name not in workbook.sheetnames:
        return result

    ws = workbook[sheet_name]

    for row in ws.iter_rows(min_row=2, values_only=True):

        key = row[0]
        value = row[1] if len(row) > 1 else None

        if key:
            result[str(key).strip()] = value

    return result


def headers_for(ws):

    headers = {}

    for index, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = index

    return headers


def cell_value(row_values, headers, column_name):

    column_index = headers.get(column_name)

    if not column_index:
        return None

    return row_values[column_index - 1]


def create_alarm():
    return [{
        "mode": "AboveValue",
        "setpointA": 0,
        "activePipeline": "Backend/RemoteNotify",
        "shelvingAllowed": False,
        "name": "Alarm",
        "priority": "High",
        "inclusiveA": False
    }]


def main():
    root_window = tk.Tk()
    root_window.lift()
    root_window.attributes('-topmost', True)
    root_window.withdraw()

    print("Opening file dialog...")
    workbook_file = filedialog.askopenfilename(
        title="Select Ignition Tag Workbook",
        filetypes=[
            ("Excel Files", "*.xlsx *.xlsm"),
            ("All Files", "*.*")
        ]
    )
    print("Workbook selected:", workbook_file)

    if not workbook_file:
        return

    output_file = filedialog.asksaveasfilename(
        title="Save Ignition JSON",
        defaultextension=".json",
        initialfile="IgnitionTags.json",
        filetypes=[("JSON Files", "*.json")]
    )

    if not output_file:
        return

    try:

        wb = load_workbook(
            workbook_file,
            data_only=True
        )

        settings = read_key_value_sheet(
            wb,
            "Settings"
        )

        ws = wb["Tags"]

        headers = headers_for(ws)

        required_headers = [
            "include",
            "name",
            "dataType",
            "documentation"
        ]

        missing = [
            x for x in required_headers
            if x not in headers
        ]

        if missing:
            raise Exception(
                "Missing required columns:\\n\\n{}".format(
                    "\\n".join(missing)
                )
            )

        root_folder = (
            text(settings.get("Root Folder"))
            or "Imported_Tags"
        )

        opc_server = (
            text(settings.get("OPC Server Default"))
            or "Ignition OPC UA Server"
        )

        opc_pattern = (
            text(settings.get("OPC Item Path Pattern"))
            or "ns=1;s=[PLC]{TagName}"
        )

        errors = []
        seen_paths = set()
        tags_by_folder = {}

        for excel_row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2
        ):

            include = bool_from_excel(
                cell_value(row, headers, "include"),
                True
            )

            if not include:
                continue

            name = text(
                cell_value(row, headers, "name")
            )

            folder_path = text(
                cell_value(row, headers, "folderPath")
            )

            datatype = text(
                cell_value(row, headers, "dataType")
            ).upper()

            documentation = text(
                cell_value(row, headers, "documentation")
            )

            eng_unit = text(
                cell_value(row, headers, "engUnit")
            )

            eng_low = cell_value(
                row, headers, "engLow"
            )

            eng_high = cell_value(
                row, headers, "engHigh"
            )

            historize = bool_from_excel(
                cell_value(row, headers, "Historize")
            )

            alarm = bool_from_excel(
                cell_value(row, headers, "alarm")
            )

            if not name:
                errors.append(
                    "Row {}: Missing tag name".format(
                        excel_row_num
                    )
                )
                continue

            if datatype not in DATATYPE_MAP:
                errors.append(
                    "Row {}: Invalid datatype '{}'".format(
                        excel_row_num,
                        datatype
                    )
                )
                continue

            full_path = "{}/{}".format(
                folder_path,
                name
            )

            if full_path in seen_paths:
                errors.append(
                    "Duplicate tag path: {}".format(
                        full_path
                    )
                )
                continue

            seen_paths.add(full_path)

            if (
                eng_low not in (None, "")
                and eng_high not in (None, "")
            ):
                try:
                    if float(eng_low) >= float(eng_high):
                        errors.append(
                            "Row {}: engLow >= engHigh".format(
                                excel_row_num
                            )
                        )
                except Exception:
                    errors.append(
                        "Row {}: Invalid engineering limits".format(
                            excel_row_num
                        )
                    )

            tag = {
                "name": name,
                "tagType": "AtomicTag",
                "dataType": DATATYPE_MAP[datatype],
                "valueSource": "opc",
                "opcServer": opc_server,
                "opcItemPath": opc_pattern.replace(
                    "{TagName}",
                    name
                ),
                "documentation": documentation
            }

            if eng_unit:
                tag["engUnit"] = eng_unit

            if eng_low not in (None, ""):
                tag["engLow"] = float(eng_low)

            if eng_high not in (None, ""):
                tag["engHigh"] = float(eng_high)

            if historize:
                tag.update({
                    "historyProvider": "TagHistorian",
                    "historyMaxAgeUnits": "SEC",
                    "historyEnabled": True,
                    "deadbandMode": "Absolute",
                    "sampleMode": "Periodic",
                    "historySampleRate": 1
                })

            if alarm:
                tag["alarms"] = create_alarm()

            tags_by_folder.setdefault(
                folder_path,
                []
            ).append(tag)

        if errors:
            raise Exception("\\n".join(errors))

        root = {
            "name": root_folder,
            "tagType": "Folder",
            "tags": []
        }

        folder_nodes = {"": root}

        def get_folder_node(path):

            path = text(path).strip("/")

            if path in folder_nodes:
                return folder_nodes[path]

            parent_path = "/".join(
                path.split("/")[:-1]
            )

            folder_name = path.split("/")[-1]

            parent = get_folder_node(parent_path)

            node = {
                "name": folder_name,
                "tagType": "Folder",
                "tags": []
            }

            parent["tags"].append(node)

            folder_nodes[path] = node

            return node

        for folder_path, tag_list in tags_by_folder.items():
            node = get_folder_node(folder_path)
            node["tags"].extend(tag_list)

        with open(
            output_file,
            "w",
            encoding="utf-8"
        ) as fp:

            json.dump(
                root,
                fp,
                indent=2
            )

        messagebox.showinfo(
            "Success",
            "Created JSON file:\\n\\n{}".format(
                output_file
            )
        )

    except Exception as exc:

        messagebox.showerror(
            "Export Failed",
            str(exc)
        )


if __name__ == "__main__":
    main()
